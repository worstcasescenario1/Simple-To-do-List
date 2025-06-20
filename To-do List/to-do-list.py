import os
from openpyxl import load_workbook

# Get current file's directory
base_path = os.path.dirname(os.path.abspath(__file__))
file_name = os.path.join(base_path, "to-do-list-dtbs.xlsx")

# Load the existing workbook
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    sheet = wb["Sheet1"]  # Make sure this matches your actual sheet name
else:
    print("File not found! Make sure it's named 'to-do-list-dtbs.xlsx'")
    exit()
# Step 2: User input loop
while True:
    task = input("Enter To-do Task (or type 'exit' to quit): ").strip()
    if task.lower() == 'exit':
        break

    sheet.append([task])
    print(f"Task added: {task}")

    while True:
        add_input = input(
            "Do you want to add another task? (Y/N): ").strip().lower()
        if add_input in ['n', 'exit']:
            print("Saving your to-do list...")
            wb.save(file_name)
            print("Done! Your list is saved at:",
                  os.path.abspath(file_name))
            print("📋 Final To-Do List:")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print("-", row[0])
            exit()
        elif add_input == 'y':
            break
        else:
            print("Please enter Y or N only.")

# Final save
wb.save(file_name)
print("✔️ To-do list saved to", os.path.abspath(file_name))
